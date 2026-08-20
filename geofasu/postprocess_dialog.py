# -*- coding: utf-8 -*-
from qgis.PyQt import uic
from qgis.PyQt.QtWidgets import (
    QDialog,
    QFileDialog,
    QMessageBox,
)
from qgis.PyQt.QtCore import Qt, QUrl
import os
import glob
import bz2
import re
import subprocess
import time
import ctypes
from ctypes import wintypes
from pathlib import Path

from qgis.core import (
    QgsProject,
    QgsProcessingContext,
    QgsProcessingFeedback,
    QgsVectorLayer,
)
import processing

FORM_CLASS, _ = uic.loadUiType(
    os.path.join(os.path.dirname(__file__), 'postprocess_dialog.ui')
)

# =========================================================
# IMPORT SCRIPTS MODULES
# =========================================================

from .scripts.psu_master_table import create_psu_master_table
from .scripts.merge_geoms import merge_geoms
from .scripts.lfs_geoid import add_lfs_geoid
from .scripts.join_validation import get_unjoinable_psu
from .scripts.load_style_samples import load_style
from .scripts.extract_missing_update import extract_missing_update
from .scripts.refactor import build_refactor_params
GEOID_COLUMN_INDEX = 0  # Excel column A


from .scripts.sensitive_loader import (
    get_csdbe_merge_module,
    get_csdbe_export_module,
)


_csdbe_merge = get_csdbe_merge_module()
_csdbe_export = get_csdbe_export_module()


build_csdbe_password_candidates = (
    _csdbe_merge.build_csdbe_password_candidates
)

build_expected_csdbe_names = (
    _csdbe_merge.build_expected_csdbe_names
)

find_csconcat_executable = (
    _csdbe_merge.find_csconcat_executable
)

merge_csdbe_files_with_password_fallback = (
    _csdbe_merge.merge_csdbe_files_with_password_fallback
)

open_csdbe_for_dictionary_export = (
    _csdbe_merge.open_csdbe_for_dictionary_export
)

validate_csdbe_files = (
    _csdbe_merge.validate_csdbe_files
)


export_csdbe_to_csv = (
    _csdbe_export.export_csdbe_to_csv
)

find_csexport_executable = (
    _csdbe_export.find_csexport_executable
)

from .scripts.household_update_validation import (
    validate_household_update_codes,
)


class PostProcessDialog(QDialog, FORM_CLASS):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

        # CADCS.pen uses a fixed path per rollout. The user copies the
        # current rollout CADCS.pen to the rollout root; this field only
        # displays the expected location.
        self.line_cadcs_path.setReadOnly(True)

        # Production mode: the former "require all CSDBE files" option is
        # obsolete. CSDBE processing remains optional, but when enabled every
        # expected CSDBE file is mandatory. Remove the old checkbox if it is
        # still present in an older copy of postprocess_dialog.ui.
        if hasattr(self, "check_require_all_csdbe"):
            obsolete_checkbox = self.check_require_all_csdbe
            obsolete_checkbox.hide()
            obsolete_checkbox.setParent(None)
            obsolete_checkbox.deleteLater()

        self.btn_browse_excel.clicked.connect(self.browse_excel)
        self.btn_read_excel.clicked.connect(self.read_excel)
        self.btn_browse_validated.clicked.connect(self.browse_validated)
        self.btn_browse_output.clicked.connect(self.browse_output)
        self.btn_browse_csdbe_data.clicked.connect(self.browse_csdbe_data)
        self.btn_browse_dictionary.clicked.connect(self.browse_dictionary)
        self.btn_browse_csconcat.clicked.connect(self.browse_csconcat)
        self.btn_process_csdbe.clicked.connect(self.process_csdbe)
        self.btn_load_all.clicked.connect(self.load_all)
        self.check_process_csdbe.toggled.connect(self.on_csdbe_option_changed)

        self.excel_rows = []
        self.validated_path = ""
        self.output_path = ""
        self.csdbe_data_path = ""
        self.cadcs_pen_path = ""
        self.detected_gpkg_files = []

        self.sample_year = None
        self.sample_month_number = None
        self.sample_month_name = ""
        self.expected_csdbe_names = []
        self.csdbe_merge_completed = False
        self.merged_csdbe_path = ""
        self.exported_csdbe_csv_path = ""
        self.exported_csdbe_csv_rows = 0

        self._set_default_csconcat_path()
        self._set_default_dictionary_path()

        # CSDBE details are hidden until the user enables CSDBE processing.
        self._set_csdbe_details_visible(
            self.check_process_csdbe.isChecked()
        )

        self._update_action_state()

    # =========================================================
    # DEFAULT PATHS
    # =========================================================

    def _set_default_csconcat_path(self):
        try:
            self.line_csconcat_path.setText(str(find_csconcat_executable()))
        except FileNotFoundError:
            # Keep blank; the user can browse manually.
            self.line_csconcat_path.clear()

    def _set_default_dictionary_path(self):
        possible_dictionary = os.path.join(
            os.path.dirname(__file__),
            "cspro",
            "LFS.dcf"
        )
        if os.path.isfile(possible_dictionary):
            self.line_csdbe_dictionary.setText(possible_dictionary)

    def create_runtime_csdbe_export_specification(
        self,
        dictionary_path,
        output_folder,
    ):
        """
        Create a run-specific CSPro export specification in the current
        GEOMS output folder. The static template stays inside the plugin.
        """
        plugin_folder = os.path.dirname(__file__)

        template_path = os.path.join(
            plugin_folder,
            "cspro",
            "LFS_HOUSEHOLD_EXPORT_TEMPLATE.exf",
        )

        if not os.path.isfile(template_path):
            raise FileNotFoundError(
                "The bundled CSPro export template was not found.\n\n"
                f"Expected template:\n{template_path}\n\n"
                "Place LFS_HOUSEHOLD_EXPORT_TEMPLATE.exf inside the "
                "plugin's cspro folder."
            )

        dictionary_path = os.path.abspath(
            str(dictionary_path or "").strip()
        )

        if not dictionary_path or not os.path.isfile(dictionary_path):
            raise FileNotFoundError(
                "The exported CSPro dictionary was not found.\n\n"
                f"Dictionary:\n{dictionary_path or '(blank)'}"
            )

        output_folder = os.path.abspath(
            str(output_folder or "").strip()
        )

        if not output_folder:
            raise RuntimeError(
                "The GEOMS output folder is blank, so the runtime EXF "
                "cannot be created."
            )

        os.makedirs(output_folder, exist_ok=True)

        try:
            with open(
                template_path,
                "r",
                encoding="utf-8-sig",
            ) as template_file:
                template_content = template_file.read()
        except OSError as exc:
            raise RuntimeError(
                "The bundled CSPro export template could not be read.\n\n"
                f"Template:\n{template_path}\n\n"
                f"Error:\n{exc}"
            ) from exc

        placeholder = "{{DICTIONARY_FILE}}"

        if placeholder not in template_content:
            raise RuntimeError(
                "The bundled CSPro export template is invalid because it "
                f"does not contain the placeholder {placeholder}.\n\n"
                f"Template:\n{template_path}"
            )

        cspro_dictionary_path = dictionary_path.replace("/", "\\")

        runtime_content = template_content.replace(
            placeholder,
            cspro_dictionary_path,
        )

        runtime_exf_path = os.path.join(
            output_folder,
            "LFS_HOUSEHOLD_EXPORT.exf",
        )

        temporary_path = runtime_exf_path + ".tmp"

        try:
            with open(
                temporary_path,
                "w",
                encoding="utf-8",
                newline="\n",
            ) as runtime_file:
                runtime_file.write(runtime_content)

            os.replace(temporary_path, runtime_exf_path)

        except OSError as exc:
            try:
                if os.path.isfile(temporary_path):
                    os.remove(temporary_path)
            except OSError:
                pass

            raise RuntimeError(
                "The runtime CSPro export specification could not be "
                "created.\n\n"
                f"Output EXF:\n{runtime_exf_path}\n\n"
                f"Error:\n{exc}"
            ) from exc

        return runtime_exf_path

    def load_csdbe_csv_to_data_group(self, csv_path, data_group):
        """
        Load the exported CSDBE household CSV as a non-spatial table
        and place it directly under the supplied DATA group.

        The CSDBE export is expected to contain the field UPDCODE.
        """
        csv_path = os.path.abspath(str(csv_path or "").strip())

        if not csv_path:
            raise RuntimeError("The exported CSDBE CSV path is empty.")

        if not os.path.isfile(csv_path):
            raise FileNotFoundError(
                f"Exported CSDBE household CSV was not found:\n{csv_path}"
            )

        csv_uri = (
            QUrl.fromLocalFile(csv_path).toString()
            + "?type=csv"
            + "&delimiter=,"
            + "&encoding=UTF-8"
            + "&detectTypes=yes"
            + "&geomType=none"
            + "&trimFields=yes"
            + "&skipEmptyFields=yes"
            + "&watchFile=no"
        )

        csv_layer = QgsVectorLayer(
            csv_uri,
            "CSDBE HOUSEHOLDS",
            "delimitedtext"
        )

        if not csv_layer.isValid():
            raise RuntimeError(
                "QGIS could not load the exported CSDBE household CSV.\n"
                f"CSV file:\n{csv_path}"
            )

        field_lookup = {
            field.name().strip().upper(): field.name()
            for field in csv_layer.fields()
        }

        if "UPDCODE" not in field_lookup:
            available_fields = ", ".join(
                field.name() for field in csv_layer.fields()
            ) or "(none)"

            raise RuntimeError(
                "The exported CSDBE household CSV does not contain the "
                "required UPDCODE field.\n\n"
                f"Available fields:\n{available_fields}"
            )

        project = QgsProject.instance()

        # Register the table without placing it at the root, then insert it
        # directly into DATA so it is visible in the Layers panel.
        project.addMapLayer(csv_layer, False)
        data_group.addLayer(csv_layer)

        return csv_layer

    # =========================================================
    # BROWSE FUNCTIONS
    # =========================================================

    def browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel file",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self.excel_path.setText(path)
            self._reset_after_input_change()

    def browse_validated(self):
        path = QFileDialog.getExistingDirectory(
            self,
            "Select Validated Path"
        )
        if path:
            self.validated_path = path
            self.line_validated_path.setText(path)
            self.csdbe_data_path = os.path.join(path, "DATA")
            self.line_csdbe_data_path.setText(self.csdbe_data_path)
            self._detect_geopackages()
            self._reset_csdbe_merge_state()
            self._update_action_state()

    def browse_output(self):
        path = QFileDialog.getExistingDirectory(
            self,
            "Select Output Path"
        )
        if path:
            self.output_path = path
            self.line_output_path.setText(path)
            self._reset_csdbe_merge_state()
            self._update_action_state()

    def browse_csdbe_data(self):
        start_path = self.line_csdbe_data_path.text().strip() or self.validated_path
        path = QFileDialog.getExistingDirectory(
            self,
            "Select CSDBE DATA Folder",
            start_path
        )
        if path:
            self.csdbe_data_path = path
            self.line_csdbe_data_path.setText(path)
            self._reset_csdbe_merge_state()
            self._update_action_state()

    def browse_dictionary(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select LFS CSPro Dictionary",
            "",
            "CSPro Dictionary (*.dcf);;All Files (*.*)"
        )
        if path:
            self.line_csdbe_dictionary.setText(path)
            self._reset_csdbe_merge_state()
            self._update_action_state()

    def browse_csconcat(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select CSPro 7.7 CSConcat.exe",
            r"C:\Program Files (x86)\CSPro 7.7",
            "CSConcat (CSConcat.exe);;Executable Files (*.exe)"
        )
        if path:
            self.line_csconcat_path.setText(path)
            self._reset_csdbe_merge_state()
            self._update_action_state()


    def _refresh_cadcs_status(self):
        """
        Refresh the fixed rollout CADCS.pen path and visual status.

        Expected location:
            <rollout root>\\CADCS.pen
        """
        cadcs_path = str(self.cadcs_pen_path or "").strip()
        self.line_cadcs_path.setText(cadcs_path)

        if cadcs_path and os.path.isfile(cadcs_path):
            self.line_cadcs_path.setStyleSheet(
                "QLineEdit {"
                "background-color:#ecfdf3;"
                "border:1px solid #4caf6a;"
                "border-radius:6px;"
                "padding:6px 9px;"
                "color:#14532d;"
                "font-weight:600;"
                "}"
            )
            self.line_cadcs_path.setToolTip(
                "CADCS.pen found for the current rollout."
            )
            if hasattr(self, "label_cadcs_instruction"):
                self.label_cadcs_instruction.setText("✓ CADCS.pen found")
                self.label_cadcs_instruction.setStyleSheet(
                    "color:#15803d; font-weight:600;"
                )
        else:
            self.line_cadcs_path.setStyleSheet(
                "QLineEdit {"
                "background-color:#fff7ed;"
                "border:1px solid #f59e0b;"
                "border-radius:6px;"
                "padding:6px 9px;"
                "color:#9a3412;"
                "}"
            )
            self.line_cadcs_path.setToolTip(
                "Copy the current rollout CADCS.pen to the displayed path."
            )
            if hasattr(self, "label_cadcs_instruction"):
                self.label_cadcs_instruction.setText(
                    "Copy CADCS.pen to this rollout folder"
                )
                self.label_cadcs_instruction.setStyleSheet(
                    "color:#b45309; font-weight:600;"
                )

    def _extract_rollout_password_candidates(self, cadcs_pen_path):
        """
        Read authorized CSDBE connection credentials from the selected rollout
        CADCS.pen.

        The CADCS package stores connection strings inside a BZip2 payload.
        Depending on rollout/compiler details, readable strings can be UTF-16LE
        or single-byte text, so both representations are inspected.
        """
        pen = Path(cadcs_pen_path)

        if not pen.is_file():
            raise FileNotFoundError(
                f"Selected CADCS.pen was not found:\n{pen}"
            )

        try:
            raw = pen.read_bytes()
        except OSError as exc:
            raise RuntimeError(
                "The selected CADCS.pen could not be read."
            ) from exc

        try:
            decoded = bz2.decompress(raw)
        except (OSError, EOFError, ValueError) as exc:
            raise RuntimeError(
                "The selected CADCS.pen could not be decoded as a supported "
                "CSPro rollout package.\n\n"
                "Select the CADCS.pen supplied with the current rollout."
            ) from exc

        # NOTE: these are real byte ranges. Do not double-escape them.
        utf16_string_re = re.compile(
            rb"(?:[\x20-\x7e]\x00){4,}"
        )
        ascii_string_re = re.compile(
            rb"[\x20-\x7e]{6,}"
        )

        # Construct the sensitive marker without showing it in UI text.
        marker = "|" + "pass" + "word="
        credential_re = re.compile(
            re.escape(marker) + r"([^|\s;\x00\r\n]+)",
            re.IGNORECASE,
        )

        found = []
        seen = set()

        def collect_from_text(readable_text):
            for credential_match in credential_re.finditer(readable_text):
                candidate = credential_match.group(1).strip()

                if not candidate:
                    continue

                key = candidate.casefold()

                if key in seen:
                    continue

                seen.add(key)
                found.append(candidate)

        for string_match in utf16_string_re.finditer(decoded):
            try:
                collect_from_text(
                    string_match.group().decode("utf-16le")
                )
            except UnicodeDecodeError:
                continue

        for string_match in ascii_string_re.finditer(decoded):
            try:
                collect_from_text(
                    string_match.group().decode("latin-1")
                )
            except UnicodeDecodeError:
                continue

        if not found:
            raise RuntimeError(
                "CADCS.pen was decoded successfully, but no CSDBE connection "
                "credential was found inside it.\n\n"
                "This rollout may use a different CADCS package layout. "
                "The DATA files were not modified."
            )

        # Prefer credentials containing the current YYMM token. CADCS can
        # contain old/legacy connection templates, so all other discovered
        # candidates remain as fallbacks.
        yymm = (
            f"{int(self.sample_year) % 100:02d}"
            f"{int(self.sample_month_number):02d}"
        )

        return sorted(
            found,
            key=lambda value: (
                yymm.casefold() not in value.casefold(),
                len(value),
            ),
        )

    def _password_candidates_for_rollout(self):
        cadcs_path = str(self.cadcs_pen_path or "").strip()

        if not cadcs_path:
            raise RuntimeError(
                "The rollout CADCS.pen path has not been determined. "
                "Read the sample workbook first."
            )

        if not os.path.isfile(cadcs_path):
            raise FileNotFoundError(
                "CADCS.pen was not found for the current rollout.\n\n"
                "Copy the current rollout CADCS.pen to:\n"
                f"{cadcs_path}"
            )

        extracted = self._extract_rollout_password_candidates(cadcs_path)
        fallback = build_csdbe_password_candidates(
            self.sample_month_number,
            self.sample_year
        )
        return list(dict.fromkeys(extracted + list(fallback)))

    @staticmethod
    def _find_dataviewer_executable():
        candidates = []
        for env_name in ("ProgramFiles(x86)", "ProgramFiles"):
            base = os.environ.get(env_name)
            if base:
                candidates.append(Path(base) / "CSPro 7.7" / "DataViewer.exe")
        candidates.extend([
            Path(r"C:\Program Files (x86)\CSPro 7.7\DataViewer.exe"),
            Path(r"C:\Program Files\CSPro 7.7\DataViewer.exe"),
        ])
        for candidate in candidates:
            if candidate.is_file():
                return candidate
        raise FileNotFoundError("CSPro 7.7 DataViewer.exe was not found on this computer.")

    @staticmethod
    def _window_text(hwnd):
        if os.name != "nt" or not hwnd:
            return ""
        user32 = ctypes.windll.user32
        length = user32.GetWindowTextLengthW(hwnd)
        if length <= 0:
            return ""
        buffer = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(hwnd, buffer, length + 1)
        return buffer.value

    @staticmethod
    def _class_name(hwnd):
        if os.name != "nt" or not hwnd:
            return ""
        buffer = ctypes.create_unicode_buffer(256)
        ctypes.windll.user32.GetClassNameW(hwnd, buffer, len(buffer))
        return buffer.value

    @classmethod
    def _wait_for_process_window(cls, pid, timeout=12.0):
        """Wait for the main visible window belonging to a process."""
        if os.name != "nt":
            return None

        user32 = ctypes.windll.user32
        found = {"hwnd": None}
        WNDENUMPROC = ctypes.WINFUNCTYPE(
            wintypes.BOOL,
            wintypes.HWND,
            wintypes.LPARAM,
        )

        def callback(hwnd, _):
            process_id = wintypes.DWORD()
            user32.GetWindowThreadProcessId(
                hwnd,
                ctypes.byref(process_id),
            )

            if (
                process_id.value == pid
                and user32.IsWindowVisible(hwnd)
                and cls._class_name(hwnd) != "#32770"
            ):
                found["hwnd"] = hwnd
                return False

            return True

        deadline = time.time() + timeout

        while time.time() < deadline:
            found["hwnd"] = None
            user32.EnumWindows(
                WNDENUMPROC(callback),
                0,
            )

            if found["hwnd"]:
                return found["hwnd"]

            time.sleep(0.20)

        return None

    @staticmethod
    def _clean_menu_text(value):
        text = str(value or "").replace("&", "").strip()

        if "\t" in text:
            text = text.split("\t", 1)[0].strip()

        text = text.replace("…", "").strip()

        while text.endswith("."):
            text = text[:-1].rstrip()

        return text.casefold()

    @classmethod
    def _menu_item(cls, menu_handle, wanted_text):
        """
        Find a menu item by its visible English label.

        Returns:
            (position, command_id, submenu_handle, visible_text)
        """
        if not menu_handle:
            return None

        user32 = ctypes.windll.user32
        MF_BYPOSITION = 0x00000400
        count = user32.GetMenuItemCount(menu_handle)
        wanted = cls._clean_menu_text(wanted_text)

        for position in range(max(0, count)):
            buffer = ctypes.create_unicode_buffer(512)

            user32.GetMenuStringW(
                menu_handle,
                position,
                buffer,
                len(buffer),
                MF_BYPOSITION,
            )

            visible_text = buffer.value
            cleaned = cls._clean_menu_text(visible_text)

            if cleaned == wanted:
                command_id = user32.GetMenuItemID(
                    menu_handle,
                    position,
                )

                submenu = user32.GetSubMenu(
                    menu_handle,
                    position,
                )

                return (
                    position,
                    command_id,
                    submenu,
                    visible_text,
                )

        return None

    @classmethod
    def _invoke_dictionary_save_menu(cls, main_hwnd):
        """
        Invoke Data Viewer:
            File -> Save As -> Dictionary

        This uses the Windows menu command IDs directly. It does not depend
        on keyboard focus or simulated Alt-key sequences.
        """
        if os.name != "nt":
            raise RuntimeError(
                "Automatic dictionary extraction is supported on Windows only."
            )

        user32 = ctypes.windll.user32
        WM_COMMAND = 0x0111

        main_menu = user32.GetMenu(main_hwnd)

        if not main_menu:
            raise RuntimeError(
                "CSPro Data Viewer opened, but its main menu could not be read."
            )

        file_item = cls._menu_item(
            main_menu,
            "File",
        )

        if not file_item or not file_item[2]:
            raise RuntimeError(
                "CSPro Data Viewer File menu was not found."
            )

        file_menu = file_item[2]

        save_as_item = cls._menu_item(
            file_menu,
            "Save As",
        )

        if not save_as_item:
            raise RuntimeError(
                "CSPro Data Viewer File -> Save As menu was not found."
            )

        dictionary_command = None

        # In CSPro 7.7 Save As is normally a submenu.
        if save_as_item[2]:
            dictionary_command = cls._menu_item(
                save_as_item[2],
                "Dictionary",
            )

        # Defensive fallback: some builds may expose Dictionary directly.
        if dictionary_command is None:
            dictionary_command = cls._menu_item(
                file_menu,
                "Dictionary",
            )

        if not dictionary_command:
            raise RuntimeError(
                "CSPro Data Viewer Save As -> Dictionary command was not found."
            )

        command_id = dictionary_command[1]

        # A submenu entry should have a concrete command identifier.
        if command_id in (-1, 0xFFFFFFFF):
            raise RuntimeError(
                "CSPro Data Viewer returned an invalid Dictionary command ID."
            )

        user32.SendMessageW(
            main_hwnd,
            WM_COMMAND,
            command_id,
            0,
        )

    @classmethod
    def _wait_for_dialog(cls, pid, owner_hwnd=None, timeout=10.0):
        """
        Wait for a visible Windows dialog (#32770) belonging to Data Viewer.
        """
        if os.name != "nt":
            return None

        user32 = ctypes.windll.user32
        found = {"hwnd": None}
        WNDENUMPROC = ctypes.WINFUNCTYPE(
            wintypes.BOOL,
            wintypes.HWND,
            wintypes.LPARAM,
        )

        def callback(hwnd, _):
            process_id = wintypes.DWORD()
            user32.GetWindowThreadProcessId(
                hwnd,
                ctypes.byref(process_id),
            )

            if process_id.value != pid:
                return True

            if not user32.IsWindowVisible(hwnd):
                return True

            if cls._class_name(hwnd) != "#32770":
                return True

            # Prefer an owned dialog when available.
            if owner_hwnd:
                owner = user32.GetWindow(
                    hwnd,
                    4,  # GW_OWNER
                )

                if owner not in (0, owner_hwnd):
                    return True

            found["hwnd"] = hwnd
            return False

        deadline = time.time() + timeout

        while time.time() < deadline:
            found["hwnd"] = None
            user32.EnumWindows(
                WNDENUMPROC(callback),
                0,
            )

            if found["hwnd"]:
                return found["hwnd"]

            time.sleep(0.20)

        return None

    @staticmethod
    def _descendant_windows(parent_hwnd):
        if os.name != "nt" or not parent_hwnd:
            return []

        user32 = ctypes.windll.user32
        result = []

        WNDENUMPROC = ctypes.WINFUNCTYPE(
            wintypes.BOOL,
            wintypes.HWND,
            wintypes.LPARAM,
        )

        def callback(hwnd, _):
            result.append(hwnd)
            return True

        user32.EnumChildWindows(
            parent_hwnd,
            WNDENUMPROC(callback),
            0,
        )

        return result

    @classmethod
    def _set_save_dialog_path(cls, dialog_hwnd, target_path):
        """
        Put the exact DCF path into the Windows Save As dialog.

        The standard File name combo has control ID 1148. Different Windows
        builds wrap its edit control differently, so descendant Edit controls
        are also inspected.
        """
        if os.name != "nt":
            return False

        user32 = ctypes.windll.user32
        WM_SETTEXT = 0x000C
        target_text = str(target_path)

        file_name_control = user32.GetDlgItem(
            dialog_hwnd,
            1148,  # cmb13 / File name
        )

        if file_name_control:
            descendants = cls._descendant_windows(
                file_name_control
            )

            for hwnd in descendants:
                if cls._class_name(hwnd).casefold() == "edit":
                    user32.SendMessageW(
                        hwnd,
                        WM_SETTEXT,
                        0,
                        target_text,
                    )
                    return True

            # Some common-dialog variants accept WM_SETTEXT on the combo.
            user32.SendMessageW(
                file_name_control,
                WM_SETTEXT,
                0,
                target_text,
            )
            return True

        # Fallback: locate an Edit whose ancestor is control 1148.
        for hwnd in cls._descendant_windows(dialog_hwnd):
            if cls._class_name(hwnd).casefold() != "edit":
                continue

            parent = user32.GetParent(hwnd)

            while parent and parent != dialog_hwnd:
                if user32.GetDlgCtrlID(parent) == 1148:
                    user32.SendMessageW(
                        hwnd,
                        WM_SETTEXT,
                        0,
                        target_text,
                    )
                    return True

                parent = user32.GetParent(parent)

        return False

    @staticmethod
    def _click_dialog_button(dialog_hwnd, control_id):
        if os.name != "nt":
            return False

        user32 = ctypes.windll.user32
        BM_CLICK = 0x00F5

        button = user32.GetDlgItem(
            dialog_hwnd,
            control_id,
        )

        if not button:
            return False

        user32.SendMessageW(
            button,
            BM_CLICK,
            0,
            0,
        )

        return True

    @classmethod
    def _dismiss_overwrite_dialog_if_present(
        cls,
        pid,
        owner_hwnd,
        timeout=1.5,
    ):
        """
        Defensive handler for a rare overwrite/confirmation prompt.
        The target file is normally deleted before export, so this is usually
        not reached.
        """
        dialog = cls._wait_for_dialog(
            pid,
            owner_hwnd=owner_hwnd,
            timeout=timeout,
        )

        if not dialog:
            return

        title = cls._window_text(dialog).casefold()

        if (
            "confirm" in title
            or "save as" in title
            or "replace" in title
        ):
            # IDYES = 6; fall back to IDOK = 1.
            if not cls._click_dialog_button(dialog, 6):
                cls._click_dialog_button(dialog, 1)

    def _automate_dictionary_export(
        self,
        source_file,
        password,
        dictionary_path,
    ):
        """
        Fully unattended CSPro 7.7 embedded-dictionary export.

        The data source is opened with its rollout credential in the connection
        string. Data Viewer's menu command and Windows Save As dialog are then
        driven directly using Win32 APIs. No user keyboard/mouse interaction
        and no foreground-window focus are required.
        """
        if os.name != "nt":
            raise RuntimeError(
                "Automatic dictionary extraction is supported on Windows only."
            )

        viewer = self._find_dataviewer_executable()
        target = Path(dictionary_path)
        source = Path(source_file)

        if not source.is_file():
            raise FileNotFoundError(
                f"CSDBE dictionary source file was not found:\n{source}"
            )

        target.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        if target.exists():
            try:
                target.unlink()
            except OSError as exc:
                raise RuntimeError(
                    "The previous dictionary file could not be removed.\n\n"
                    f"File:\n{target}\n\n"
                    "Close CSPro Dictionary Designer/Data Viewer if the file "
                    "is currently open."
                ) from exc

        connection = (
            f"{source.resolve()}"
            f"|password={password}"
        )

        proc = subprocess.Popen(
            [
                str(viewer),
                connection,
            ],
            close_fds=True,
        )

        try:
            main_hwnd = self._wait_for_process_window(
                proc.pid,
                timeout=15.0,
            )

            if not main_hwnd:
                raise RuntimeError(
                    "CSPro 7.7 Data Viewer did not create its main window."
                )

            # Give Data Viewer time to finish loading/decrypting the source.
            time.sleep(1.0)

            self._invoke_dictionary_save_menu(
                main_hwnd
            )

            save_dialog = self._wait_for_dialog(
                proc.pid,
                owner_hwnd=main_hwnd,
                timeout=10.0,
            )

            if not save_dialog:
                raise RuntimeError(
                    "CSPro Data Viewer did not open the Save Dictionary dialog. "
                    "The rollout credential may not match this CSDBE, or the "
                    "selected source may not contain an embedded dictionary."
                )

            if not self._set_save_dialog_path(
                save_dialog,
                target,
            ):
                raise RuntimeError(
                    "The Windows Save Dictionary dialog opened, but its File "
                    "name field could not be located."
                )

            # IDOK = 1 is the Save button for the standard common dialog.
            if not self._click_dialog_button(
                save_dialog,
                1,
            ):
                raise RuntimeError(
                    "The Windows Save Dictionary dialog opened, but the Save "
                    "button could not be activated."
                )

            self._dismiss_overwrite_dialog_if_present(
                proc.pid,
                owner_hwnd=main_hwnd,
                timeout=1.0,
            )

            deadline = time.time() + 20.0

            while time.time() < deadline:
                if (
                    target.is_file()
                    and target.stat().st_size > 0
                ):
                    # Basic CSPro 7.7 dictionary sanity check.
                    try:
                        header = target.read_text(
                            encoding="utf-8-sig",
                            errors="ignore",
                        )[:2048]
                    except OSError:
                        header = ""

                    if (
                        "[Dictionary]" in header
                        or "Version=CSPro 7.7" in header
                    ):
                        return target

                    # A non-empty file is still accepted because some
                    # installations can emit a different compatible encoding.
                    return target

                if proc.poll() is not None:
                    break

                time.sleep(0.25)

            raise RuntimeError(
                "CSPro Data Viewer completed the automatic Save Dictionary "
                "sequence, but no non-empty .dcf file was created.\n\n"
                f"Expected dictionary:\n{target}"
            )

        finally:
            try:
                if proc.poll() is None:
                    proc.terminate()
                    proc.wait(timeout=3)
            except Exception:
                try:
                    proc.kill()
                except Exception:
                    pass

    # =========================================================
    # STATE MANAGEMENT
    # =========================================================

    def _reset_after_input_change(self):
        self.excel_rows.clear()
        self.expected_csdbe_names.clear()
        self.detected_gpkg_files.clear()
        self.text_detected_gpkg.clear()
        self.text_csdbe_files.clear()
        self.sample_year = None
        self.sample_month_number = None
        self.sample_month_name = ""
        self.cadcs_pen_path = ""
        self.line_cadcs_path.clear()
        self._refresh_cadcs_status()
        self._reset_csdbe_merge_state()
        self._update_action_state()

    def _reset_csdbe_merge_state(self):
        self.csdbe_merge_completed = False
        self.merged_csdbe_path = ""
        self.exported_csdbe_csv_path = ""
        self.exported_csdbe_csv_rows = 0

        if self.check_process_csdbe.isChecked():
            self.label_csdbe_status.setText(
                "Read Excel, verify the CSDBE settings, then press Process CSDBE."
            )
        else:
            self.label_csdbe_status.setText("CSDBE processing is disabled.")

    def _set_csdbe_details_visible(self, visible):
        """
        Show/hide the CSDBE details as one block and reserve enough vertical
        space when expanded so the grid rows cannot be compressed/overlap.
        """
        if hasattr(self, "csdbe_details_widget"):
            self.csdbe_details_widget.setVisible(visible)

        if hasattr(self, "group_csdbe"):
            if visible:
                self.group_csdbe.setMinimumHeight(355)
                self.group_csdbe.setMaximumHeight(16777215)
            else:
                self.group_csdbe.setMinimumHeight(72)
                self.group_csdbe.setMaximumHeight(95)

        if hasattr(self, "csdbe_details_widget"):
            if visible:
                self.csdbe_details_widget.setMinimumHeight(285)
                self.csdbe_details_widget.setMaximumHeight(16777215)
            else:
                self.csdbe_details_widget.setMinimumHeight(0)
                self.csdbe_details_widget.setMaximumHeight(0)

        # Recalculate the dialog/layout geometry after collapsing or expanding.
        self.layout().invalidate()
        self.layout().activate()


    def on_csdbe_option_changed(self, checked):
        self._set_csdbe_details_visible(checked)
        self._reset_csdbe_merge_state()
        self.btn_process_csdbe.setEnabled(
            checked and bool(self.excel_rows)
        )
        self._update_action_state()


    def _update_action_state(self):
        has_excel = bool(self.excel_rows)
        csdbe_enabled = self.check_process_csdbe.isChecked()

        self.btn_process_csdbe.setEnabled(has_excel and csdbe_enabled)

        # Loading can proceed only after Excel was read and either:
        # 1. CSDBE processing is disabled, or
        # 2. CSDBE processing completed successfully.
        can_load = has_excel and (
            not csdbe_enabled or self.csdbe_merge_completed
        )
        self.btn_load_all.setEnabled(can_load)

    # =========================================================
    # READ EXCEL
    # =========================================================

    def read_excel(self):
        excel_path = self.excel_path.text().strip()

        if not excel_path:
            QMessageBox.warning(self, "Missing File", "Select Excel file first.")
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(excel_path, data_only=True, read_only=True)

            sheets = ["Sample SSU", "Replacement SSU"]
            self.excel_rows.clear()

            year_value = None
            round_value = None
            province_name = None

            month_names = {
                1: "JANUARY", 2: "FEBRUARY", 3: "MARCH", 4: "APRIL",
                5: "MAY", 6: "JUNE", 7: "JULY", 8: "AUGUST",
                9: "SEPTEMBER", 10: "OCTOBER", 11: "NOVEMBER", 12: "DECEMBER"
            }

            for sheet in sheets:
                if sheet not in wb.sheetnames:
                    continue

                ws = wb[sheet]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    # GEOID is stored in Excel column A (Python index 0).
                    if len(row) <= GEOID_COLUMN_INDEX or row[GEOID_COLUMN_INDEX] in (None, ''):
                        continue

                    self.excel_rows.append(list(row[:25]))

                    if year_value is None:
                        round_value = int(row[1])
                        year_value = int(row[2])
                        province_name = row[5]

            wb.close()

            if year_value is None or round_value is None:
                raise Exception("Cannot determine Year/Round from Excel.")

            if round_value not in month_names:
                raise Exception(f"Invalid month/round value in Excel: {round_value}")

            if not self.excel_rows:
                raise Exception(
                    "No sample GEOIDs were found in Sample SSU or Replacement SSU."
                )

            month_name = month_names[round_value]
            province_name = str(province_name).strip().upper()

            self.sample_year = year_value
            self.sample_month_number = round_value
            self.sample_month_name = month_name

            rollout_root = (
                rf"C:\PSA-GIS\{province_name}\GEOFASU"
                rf"\{year_value}\{month_name}"
            )

            self.validated_path = os.path.join(
                rollout_root,
                "FIELD VALIDATION OUTPUT"
            )
            self.output_path = os.path.join(
                rollout_root,
                "GEOMS"
            )
            self.csdbe_data_path = os.path.join(
                self.validated_path,
                "DATA"
            )

            # Fixed CADCS.pen location for the rollout.
            # Example:
            # C:\PSA-GIS\ROMBLON\GEOFASU\2026\JULY\CADCS.pen
            self.cadcs_pen_path = os.path.join(
                rollout_root,
                "CADCS.pen"
            )

            os.makedirs(self.output_path, exist_ok=True)

            self.line_validated_path.setText(self.validated_path)
            self.line_output_path.setText(self.output_path)
            self.line_csdbe_data_path.setText(self.csdbe_data_path)
            self._refresh_cadcs_status()

            generated_dictionary = os.path.join(
                self.output_path,
                f"LFS_DATA_DICTIONARY_{month_name}_{year_value}.dcf"
            )
            self.line_csdbe_dictionary.setText(generated_dictionary)

            self.expected_csdbe_names = build_expected_csdbe_names(
                self.excel_rows,
                geoid_index=GEOID_COLUMN_INDEX
            )

            self._detect_geopackages()
            self._reset_csdbe_merge_state()
            self._show_initial_csdbe_summary()
            self._update_action_state()

            csdbe_note = (
                "\n\nCSDBE processing is enabled. Press Process CSDBE before Load All."
                if self.check_process_csdbe.isChecked()
                else "\n\nCSDBE processing is disabled and will be skipped."
            )

            QMessageBox.information(
                self,
                "Excel Read",
                "Excel read successfully.\n"
                f"Sample rows: {len(self.excel_rows)}\n"
                f"Expected unique CSDBE files: {len(self.expected_csdbe_names)}\n"
                f"GeoPackages detected: {len(self.detected_gpkg_files)}"
                + csdbe_note
            )

        except Exception as e:
            self._reset_after_input_change()
            QMessageBox.critical(self, "Excel Read Error", str(e))

    def _detect_geopackages(self):
        self.detected_gpkg_files.clear()

        validated_path = self.line_validated_path.text().strip() or self.validated_path
        if not validated_path:
            self.text_detected_gpkg.clear()
            return

        geopackage_root = os.path.join(validated_path, "GEOPACKAGES")
        if not os.path.isdir(geopackage_root):
            self.text_detected_gpkg.setPlainText(
                "GEOPACKAGES folder not found:\n" + geopackage_root
            )
            return

        for root_folder, _, filenames in os.walk(geopackage_root):
            for fname in filenames:
                lower = fname.casefold()
                if not lower.endswith(".gpkg"):
                    continue
                if "_lfs_" not in lower:
                    continue
                if any(x in lower for x in ("_bgy", "_img_clipped")):
                    continue
                self.detected_gpkg_files.append(os.path.join(root_folder, fname))

        self.detected_gpkg_files.sort(key=str.casefold)
        if self.detected_gpkg_files:
            lines = [
                f"GeoPackage folder: {geopackage_root}",
                f"Detected: {len(self.detected_gpkg_files)}",
                "",
                *self.detected_gpkg_files,
            ]
        else:
            lines = [
                "No qualifying LFS GeoPackages detected.",
                "",
                f"Folder searched: {geopackage_root}",
            ]
        self.text_detected_gpkg.setPlainText("\n".join(lines))

    def _show_initial_csdbe_summary(self):
        if not self.expected_csdbe_names:
            self.text_csdbe_files.clear()
            return

        lines = [
            f"Expected CSDBE files: {len(self.expected_csdbe_names)}",
            f"DATA folder: {self.csdbe_data_path}",
            "",
            "Expected filenames:",
        ]
        lines.extend(self.expected_csdbe_names)
        self.text_csdbe_files.setPlainText("\n".join(lines))

    # =========================================================
    # CSDBE PROCESSING
    # =========================================================

    def process_csdbe(self):
        if not self.check_process_csdbe.isChecked():
            QMessageBox.information(
                self,
                "CSDBE Disabled",
                "Enable CSDBE processing first."
            )
            return

        if not self.excel_rows or not self.expected_csdbe_names:
            QMessageBox.warning(
                self,
                "No Excel Data",
                "Read the Excel sample list first."
            )
            return

        data_folder = self.line_csdbe_data_path.text().strip()
        output_folder = self.line_output_path.text().strip() or self.output_path
        dictionary_path = self.line_csdbe_dictionary.text().strip()
        configured_csconcat = self.line_csconcat_path.text().strip()

        if not data_folder:
            QMessageBox.warning(self, "Missing DATA Folder", "Select the CSDBE DATA folder.")
            return

        cadcs_path = str(self.cadcs_pen_path or "").strip()
        self._refresh_cadcs_status()

        if not cadcs_path:
            QMessageBox.warning(
                self,
                "CADCS.pen Path Not Set",
                "Read the sample workbook first so GEOFASU can determine "
                "the fixed CADCS.pen rollout location."
            )
            return

        if not os.path.isfile(cadcs_path):
            QMessageBox.warning(
                self,
                "CADCS.pen Required",
                "CADCS.pen is required when CSDBE processing is enabled.\n\n"
                "Copy the CADCS.pen supplied for this rollout to:\n\n"
                f"{cadcs_path}\n\n"
                "Then press Process CSDBE again."
            )
            return

        try:
            os.makedirs(output_folder, exist_ok=True)
            self.setCursor(Qt.WaitCursor)
            self.btn_process_csdbe.setEnabled(False)
            self.btn_load_all.setEnabled(False)
            self.label_csdbe_status.setText("Validating expected CSDBE files…")

            validation = validate_csdbe_files(
                data_folder,
                self.expected_csdbe_names
            )

            summary_lines = [
                f"Expected: {len(validation.expected_files)}",
                f"Found: {len(validation.found_files)}",
                f"Missing: {len(validation.missing_files)}",
                "",
            ]

            if validation.found_files:
                summary_lines.append("FOUND:")
                summary_lines.extend(str(path) for path in validation.found_files)
                summary_lines.append("")

            if validation.missing_files:
                summary_lines.append("MISSING:")
                summary_lines.extend(path.name for path in validation.missing_files)

            self.text_csdbe_files.setPlainText("\n".join(summary_lines))

            if not validation.found_files:
                raise RuntimeError(
                    "None of the expected CSDBE files were found in the DATA folder."
                )

            # Production requirement: once CSDBE processing is enabled,
            # every expected file must be present. Partial merges are never
            # permitted. CSDBE processing itself remains optional through
            # check_process_csdbe.
            if validation.missing_files:
                missing_names = "\n".join(
                    path.name for path in validation.missing_files
                )
                self.label_csdbe_status.setText(
                    "CSDBE processing stopped: required files are missing."
                )
                raise RuntimeError(
                    "CSDBE production validation failed.\n\n"
                    "CSDBE processing was enabled, so every sample in the "
                    "Excel list must have a corresponding CSDBE file.\n\n"
                    f"Expected files: {len(validation.expected_files)}\n"
                    f"Found files: {len(validation.found_files)}\n"
                    f"Missing files: {len(validation.missing_files)}\n\n"
                    f"Missing filenames:\n{missing_names}\n\n"
                    "Copy all missing CSDBE files into the DATA folder, then "
                    "run Process CSDBE again. To skip CSDBE processing "
                    "entirely, clear the Process CSDBE option."
                )

            # The dictionary is embedded in the encrypted CSDBE.
            # Dictionary export is intentionally USER-DRIVEN for reliability.
            #
            # GEOFASU opens the first valid CSDBE in CSPro 7.7 Data Viewer using
            # the current rollout credential discovered from CADCS.pen. The user
            # then saves the embedded dictionary manually through:
            #
            #     File -> Save As -> Dictionary
            #
            # After saving the .dcf to the exact path shown by GEOFASU, press
            # Process CSDBE again. The merge and CSV export then continue.
            if not dictionary_path:
                dictionary_path = os.path.join(
                    output_folder,
                    f"LFS_DATA_DICTIONARY_{self.sample_month_name}_{self.sample_year}.dcf"
                )
                self.line_csdbe_dictionary.setText(dictionary_path)

            password_candidates = self._password_candidates_for_rollout()

            if not os.path.isfile(dictionary_path):
                source_file = validation.found_files[0]

                if not password_candidates:
                    raise RuntimeError(
                        "No rollout credential could be obtained from the selected "
                        "CADCS.pen. No CSDBE file was opened."
                    )

                # The CADCS reader ranks the current rollout credential first.
                # The credential is not displayed to the user.
                rollout_password = password_candidates[0]

                viewer = self._find_dataviewer_executable()
                connection = (
                    f"{Path(source_file).resolve()}"
                    f"|password={rollout_password}"
                )

                try:
                    subprocess.Popen(
                        [str(viewer), connection],
                        close_fds=True,
                    )
                except OSError as exc:
                    raise RuntimeError(
                        "CSPro 7.7 Data Viewer could not be opened.\n\n"
                        f"Executable:\n{viewer}\n\n"
                        f"Source CSDBE:\n{source_file}\n\n"
                        f"Windows error:\n{exc}"
                    ) from exc

                self.label_csdbe_status.setText(
                    "Waiting for user to export the embedded dictionary in "
                    "CSPro 7.7 Data Viewer."
                )

                self.unsetCursor()

                QMessageBox.information(
                    self,
                    "Export LFS Dictionary",
                    "The first valid CSDBE file has been opened in "
                    "CSPro 7.7 Data Viewer using the current rollout "
                    "authentication.\n\n"
                    f"Source CSDBE:\n{source_file}\n\n"
                    "In Data Viewer, select:\n"
                    "File -> Save As -> Dictionary\n\n"
                    "Save the dictionary EXACTLY as:\n"
                    f"{dictionary_path}\n\n"
                    "After saving the .dcf file, return to GEOFASU and "
                    "press Process CSDBE again.\n\n"
                    "GEOFASU will then continue automatically with CSDBE "
                    "merge, household CSV export, and the remaining workflow."
                )
                return

            csconcat_path = str(find_csconcat_executable(configured_csconcat))
            self.line_csconcat_path.setText(csconcat_path)

            output_filename = (
                f"MERGED_CSDBE_{self.sample_month_name}_{self.sample_year}.csdbe"
            )
            merged_output = os.path.join(output_folder, output_filename)
            pff_path = os.path.join(output_folder, "merger.pff")

            self.label_csdbe_status.setText(
                f"Merging {len(validation.found_files)} CSDBE file(s) with CSPro 7.7…"
            )

            password_merge_result = merge_csdbe_files_with_password_fallback(
                csconcat_executable=csconcat_path,
                input_files=[str(path) for path in validation.found_files],
                output_file=merged_output,
                dictionary_file=dictionary_path,
                password_candidates=password_candidates,
                pff_file=pff_path,
            )

            merge_result = password_merge_result.merge_result
            successful_password = password_merge_result.successful_password

            # -------------------------------------------------
            # EXPORT SELECTED HOUSEHOLD FIELDS TO CSV
            # -------------------------------------------------
            self.label_csdbe_status.setText(
                "CSDBE merge completed. Exporting household fields to CSV…"
            )

            csexport_path = str(find_csexport_executable())

            export_specification = (
                self.create_runtime_csdbe_export_specification(
                    dictionary_path=dictionary_path,
                    output_folder=output_folder,
                )
            )

            csv_output = os.path.join(
                output_folder,
                f"LFS_HOUSEHOLD_{self.sample_month_name}_{self.sample_year}.csv"
            )

            export_pff_path = os.path.join(
                output_folder,
                "csdbe_export.pff"
            )

            export_listing_path = os.path.join(
                output_folder,
                "csdbe_export.lst"
            )

            export_result = export_csdbe_to_csv(
                csexport_executable=csexport_path,
                export_specification_file=export_specification,
                input_csdbe_file=str(merge_result.output_file),
                output_csv_file=csv_output,
                dictionary_file=dictionary_path,
                password=successful_password,
                pff_file=export_pff_path,
                listing_file=export_listing_path,
            )

            self.csdbe_merge_completed = True
            self.merged_csdbe_path = str(merge_result.output_file)
            self.exported_csdbe_csv_path = str(export_result.output_file)
            self.exported_csdbe_csv_rows = export_result.row_count

            self.label_csdbe_status.setText(
                "CSDBE processing completed successfully. "
                f"Merged {merge_result.input_count} file(s) and exported "
                f"{export_result.row_count:,} CSV row(s)."
            )
            self._update_action_state()

            password_source = "Rollout CADCS authentication"

            # Reaching this point guarantees that all expected CSDBE files
            # were present.
            missing_note = ""

            QMessageBox.information(
                self,
                "CSDBE Processing Complete",
                f"Merged input files: {merge_result.input_count}\n"
                f"Password attempt used: "
                f"{password_merge_result.password_attempt_number}\n"
                f"Password source: {password_source}\n\n"
                f"Merged CSDBE:\n{merge_result.output_file}\n\n"
                f"CSV rows exported: {export_result.row_count:,}\n"
                f"CSV output:\n{export_result.output_file}\n\n"
                "Temporary CSPro runtime files are managed automatically.\n"
                f"{missing_note}\n\n"
                "You may now press Load All (Excel + GPKG)."
            )

        except Exception as e:
            self.csdbe_merge_completed = False
            self.merged_csdbe_path = ""
            self.exported_csdbe_csv_path = ""
            self.exported_csdbe_csv_rows = 0
            self.label_csdbe_status.setText("CSDBE processing failed.")
            self._update_action_state()
            QMessageBox.critical(self, "CSDBE Processing Error", str(e))

        finally:
            self.unsetCursor()
            self._update_action_state()

    # =========================================================
    # MAIN PROCESS — EXISTING WORKFLOW
    # =========================================================

    def load_all(self):
        if not self.excel_rows:
            QMessageBox.warning(self, "No Excel Data", "Please read Excel first.")
            return

        if self.check_process_csdbe.isChecked() and not self.csdbe_merge_completed:
            QMessageBox.warning(
                self,
                "CSDBE Not Processed",
                "Process and merge the CSDBE files before loading GeoPackages."
            )
            return

        # Re-detect in case the user changed the validated path manually.
        self.validated_path = self.line_validated_path.text().strip() or self.validated_path
        self._detect_geopackages()

        if not self.detected_gpkg_files:
            QMessageBox.warning(self, "No Files", "No GeoPackages detected.")
            return

        output_path = self.line_output_path.text().strip() or self.output_path
        os.makedirs(output_path, exist_ok=True)

        try:
            context = QgsProcessingContext()
            feedback = QgsProcessingFeedback()

            def layer_is_empty(layer):
                return (layer is None) or (layer.featureCount() == 0)

            # -------------------------------------------------
            # 1. CREATE PSU MASTER TABLE
            # -------------------------------------------------
            table_layer = create_psu_master_table(self.excel_rows, output_path)

            # -------------------------------------------------
            # 2. MERGE GEOMS
            # -------------------------------------------------
            merged_layer = merge_geoms(self.detected_gpkg_files, output_path)

            # -------------------------------------------------
            # TEMP FILTER ONLY FOR EXPORT (DO NOT TOUCH MERGED)
            # -------------------------------------------------
            from qgis.core import QgsFeatureRequest

            def extract_not_null_update_codes(layer):
                expr = '\"Update Codes\" IS NOT NULL AND \"Update Codes\" != \'\''
                request = QgsFeatureRequest().setFilterExpression(expr)
                return layer.materialize(request)

            # -------------------------------------------------
            # 3. ADD LFS GEOID
            # -------------------------------------------------
            add_lfs_geoid(merged_layer)

            # -------------------------------------------------
            # 4. CSDBE HOUSEHOLD MATCHING WITH CANVAS LOGS
            # -------------------------------------------------
            update_code_validation = None
            match_log_layer = None
            not_matched_log_layer = None

            if self.check_process_csdbe.isChecked():
                csv_file = (
                    self.exported_csdbe_csv_path
                    or os.path.join(
                        output_path,
                        f"LFS_HOUSEHOLD_{self.sample_month_name}_{self.sample_year}.csv"
                    )
                )

                self.label_csdbe_status.setText(
                    "Matching household GEOIDs and UPDCODE values…"
                )

                update_code_validation = validate_household_update_codes(
                    merged_layer=merged_layer,
                    csv_file=csv_file,
                )
                match_log_layer = update_code_validation.diagnostic_layer
                not_matched_log_layer = update_code_validation.inconsistent_layer

            # -------------------------------------------------
            # 5. UNJOINABLE PSU
            # -------------------------------------------------
            unmatched_layer = get_unjoinable_psu(table_layer, merged_layer)

            # -------------------------------------------------
            # 6. MISSING UPDATE EXTRACTION
            # -------------------------------------------------
            missing_update_layer = extract_missing_update(merged_layer)

            # -------------------------------------------------
            # 7. LOAD TO QGIS
            # -------------------------------------------------
            project = QgsProject.instance()
            root = project.layerTreeRoot()

            project.addMapLayer(table_layer, False)
            project.addMapLayer(merged_layer, False)

            if unmatched_layer:
                project.addMapLayer(unmatched_layer, False)

            if missing_update_layer:
                project.addMapLayer(missing_update_layer, False)

            if match_log_layer:
                project.addMapLayer(match_log_layer, False)

            if not_matched_log_layer:
                project.addMapLayer(not_matched_log_layer, False)

            load_style(merged_layer, "merged_geoms.qml")

            if missing_update_layer:
                load_style(missing_update_layer, "samples_no_uc.qml")

            # -------------------------------------------------
            # REMOVE OLD GROUPS
            # -------------------------------------------------
            for group_name in ["INVALID FEATURES", "MERGED FEATURES", "DATA"]:
                grp = root.findGroup(group_name)
                if grp:
                    root.removeChildNode(grp)

            # -------------------------------------------------
            # CREATE GROUPS
            # -------------------------------------------------
            invalid_group = root.addGroup("INVALID FEATURES")
            merged_group = root.addGroup("MERGED FEATURES")
            data_group = root.addGroup("DATA")

            # -------------------------------------------------
            # LOAD EXPORTED CSDBE CSV INTO DATA GROUP
            # -------------------------------------------------
            csdbe_csv_layer = None

            if self.check_process_csdbe.isChecked():
                csv_file = (
                    self.exported_csdbe_csv_path
                    or os.path.join(
                        output_path,
                        f"LFS_HOUSEHOLD_{self.sample_month_name}_{self.sample_year}.csv"
                    )
                )

                csdbe_csv_layer = self.load_csdbe_csv_to_data_group(
                    csv_file,
                    data_group
                )

            # The complete log is placed in DATA for inspection.
            if match_log_layer:
                match_log_layer.setName("CSDBE MATCH LOG - ALL")
                data_group.addLayer(match_log_layer)

            # Failed matches are also placed in INVALID FEATURES.
            if not_matched_log_layer:
                not_matched_log_layer.setName("CSDBE MATCH LOG - NOT MATCHED")
                invalid_group.addLayer(not_matched_log_layer)

            # -------------------------------------------------
            # RENAME LAYERS
            # -------------------------------------------------
            merged_layer.setName("MERGED_GEOMS")
            table_layer.setName("PSU_MASTER_TABLE")

            if unmatched_layer:
                unmatched_layer.setName("MISSING SSUs")

            if missing_update_layer:
                missing_update_layer.setName("ORIGINAL SAMPLES WITH NO UPDATE CODES")

            # -------------------------------------------------
            # MOVE LAYERS INTO GROUPS
            # -------------------------------------------------
            def move_to_group(layer, group):
                node = root.findLayer(layer.id())
                if node:
                    clone = node.clone()
                    group.addChildNode(clone)
                    node.parent().removeChildNode(node)

            if missing_update_layer:
                move_to_group(missing_update_layer, invalid_group)

            if unmatched_layer:
                move_to_group(unmatched_layer, invalid_group)

            move_to_group(merged_layer, merged_group)
            move_to_group(table_layer, merged_group)

            # -------------------------------------------------
            # GROUP DISPLAY CONTROL
            # -------------------------------------------------
            invalid_group.setExpanded(True)
            merged_group.setExpanded(True)
            data_group.setExpanded(True)

            def collapse_layer(layer):
                node = root.findLayer(layer.id())
                if node:
                    node.setExpanded(False)

            for lyr in [
                merged_layer,
                table_layer,
                missing_update_layer,
                unmatched_layer,
                csdbe_csv_layer,
                match_log_layer,
                not_matched_log_layer,
            ]:
                if lyr:
                    collapse_layer(lyr)

            # -------------------------------------------------
            # REFACTOR LOGIC (UNCHANGED)
            # -------------------------------------------------
            if layer_is_empty(missing_update_layer) and layer_is_empty(unmatched_layer):
                export_layer = extract_not_null_update_codes(merged_layer)

                refactored = processing.run(
                    "native:refactorfields",
                    build_refactor_params(export_layer),
                    context=context,
                    feedback=feedback,
                    is_child_algorithm=True
                )["OUTPUT"]

                xlsx_path = os.path.join(output_path, "exported.xlsx")

                processing.run(
                    "qgis:exporttospreadsheet",
                    {
                        "LAYERS": [refactored],
                        "OUTPUT": xlsx_path
                    },
                    context=context,
                    feedback=feedback,
                    is_child_algorithm=True
                )

            if update_code_validation is not None:
                validation_summary = (
                    "Post-processing completed.\n\n"
                    "CSDBE household matching:\n"
                    f"CSV rows: {update_code_validation.csv_record_count:,}\n"
                    f"CSV unique household GEOIDs: "
                    f"{update_code_validation.csv_unique_geoid_count:,}\n"
                    f"Merged features checked: "
                    f"{update_code_validation.layer_feature_count:,}\n"
                    f"Matches: {update_code_validation.matched_count:,}\n"
                    f"Update-code mismatches: "
                    f"{update_code_validation.mismatch_count:,}\n"
                    f"Household GEOID not found in CSV: "
                    f"{update_code_validation.csv_not_found_count:,}\n"
                    f"CSV GEOID conflicts: "
                    f"{update_code_validation.csv_conflict_count:,}\n"
                    f"Invalid merged GEOIDs: "
                    f"{update_code_validation.invalid_geoid_count:,}\n\n"
                    "Open DATA > CSDBE MATCH LOG - ALL to inspect every "
                    "matching parameter and reason. Failed records are also "
                    "shown under INVALID FEATURES > "
                    "CSDBE MATCH LOG - NOT MATCHED."
                )
            else:
                validation_summary = "Post-processing completed successfully."

            QMessageBox.information(
                self,
                "Success",
                validation_summary
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))