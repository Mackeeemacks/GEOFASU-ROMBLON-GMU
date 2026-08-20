# -*- coding: utf-8 -*-
from qgis.PyQt import uic
from qgis.PyQt.QtWidgets import (
    QDialog, QFileDialog, QMessageBox, QApplication, QProgressDialog,
    QTableWidgetItem, QHeaderView
)
from qgis.PyQt.QtCore import Qt
from qgis.PyQt.QtGui import QColor, QBrush
from qgis.core import (
    QgsVectorLayer, QgsFeature, QgsGeometry, QgsPointXY,
    QgsProject, QgsProcessingFeedback, QgsVectorFileWriter, QgsCoordinateReferenceSystem,
    QgsRasterLayer, QgsSnappingConfig, QgsTolerance
)
import os, re, shutil
from pathlib import Path

# Import scripts
from .scripts.extract_psu import extract_by_psu
from .scripts.refactor_sample import refactor_psu_layer
from .scripts.load_style_samples import load_style
from .scripts.load_bgy import load_barangay_layer_smart
from .scripts.clip_raster_by_bgy import clip_raster_by_bgy_memory
from .scripts.qfield_package_inspector import inspect_current_project
from .scripts.qfield_packager import package_current_project

# Load UI
FORM_CLASS, _ = uic.loadUiType(
    os.path.join(os.path.dirname(__file__), 'geofasu_dialog_base.ui')
)


class geofasuDialog(QDialog, FORM_CLASS):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

        # --- Connect buttons ---
        self.pbBrowseExcel.clicked.connect(self.browse_excel)
        self.pbBrowseOutput.clicked.connect(self.browse_output)
        self.pbLoadSSU.clicked.connect(self.load_psu_list)
        self.pbGenerateGeometry.clicked.connect(self.generate_geometry)
        self.pbInspectQField.clicked.connect(self.inspect_qfield_project)
        self.pbPackageQField.clicked.connect(self.package_qfield_project)

        self.cbpsu_list.currentIndexChanged.connect(self.update_selected_psu_paths)
        self.project_abbreviation.textChanged.connect(
            self.on_project_abbreviation_changed
        )

        # Project abbreviation controls generated names and paths.
        self.project_abbreviation.setText("LFS")
        self.project_abbreviation.setMaxLength(20)
        self.project_abbreviation.setToolTip(
            "Short project code used in generated layer names, filenames, "
            "QGIS project names and folders. Examples: LFS, GATS, HSDV."
        )

        # Output path display-only
        self.output_path.setReadOnly(True)
        self.output_path.setToolTip(
            "The output folder is generated automatically from the selected PSU."
        )

        self.qfield_package_path.setReadOnly(True)
        self.qfield_package_path.setToolTip(
            "Default QField package folder for the currently selected PSU."
        )

        self.pbPackageQField.setEnabled(False)
        self.pbPackageQField.setToolTip(
            "Inspect the current project first. Packaging is enabled "
            "when no blocking errors are found."
        )

        self._last_qfield_inspection = None

        self.tblQFieldLayers.setColumnCount(5)
        self.tblQFieldLayers.setHorizontalHeaderLabels([
            "Layer",
            "Type",
            "Provider",
            "Proposed Action",
            "Status",
        ])
        header = self.tblQFieldLayers.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        for column in range(1, 5):
            header.setSectionResizeMode(
                column,
                QHeaderView.ResizeToContents
            )
        self.tblQFieldLayers.verticalHeader().setVisible(False)
        self.tblQFieldLayers.setAlternatingRowColors(True)
        self.tblQFieldLayers.setSelectionBehavior(
            self.tblQFieldLayers.SelectRows
        )
        self.tblQFieldLayers.setEditTriggers(
            self.tblQFieldLayers.NoEditTriggers
        )

        # --- Keep layer references to avoid deletion ---
        self.lfs_layer = None
        self.bgy_layer = None
        self.clipped_raster = None
        self.generated_project_path = None

    # =========================================================
    # Helper: Month Name
    # =========================================================
    @staticmethod
    def month_name(m):
        months = [
            "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
            "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"
        ]
        return months[m - 1] if 1 <= m <= 12 else "UNKNOWN"

    # =========================================================
    # Browse Excel
    # =========================================================
    def browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel file", "", "Excel Files (*.xlsx)"
        )
        if path:
            self.ssu_list_path.setText(path)

    # =========================================================
    # Browse Output Folder
    # =========================================================
    def browse_output(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_path.setText(folder)

    # =========================================================
    # Load PSU List
    # =========================================================
    def load_psu_list(self):
        excel_path = self.ssu_list_path.text()
        if not excel_path:
            QMessageBox.warning(self, "Missing File", "Select Excel file first.")
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            if "Sample PSU" not in wb.sheetnames:
                raise Exception("Sheet 'Sample PSU' not found.")

            ws = wb["Sample PSU"]
            self.cbpsu_list.clear()

            for row in ws.iter_rows(min_row=2, values_only=True):
                geoid, reg_name, prov_name, mun_name, psu_name, psu_number, rep, rnd, year, rg = row
                if psu_number is None or geoid is None:
                    continue

                number = int(psu_number)
                text = f"{number} ({mun_name}, {psu_name})"
                geoid_str = str(geoid)
                geoid_prefix = geoid_str[2:5] + geoid_str[5:7]

                data = {
                    "Prov_name": prov_name,
                    "PSU_number": number,
                    "Mun_name": mun_name,
                    "PSU_name": psu_name,
                    "Replicate_Number": rep,
                    "Round": rnd,
                    "Year": year,
                    "Geoid_prefix": geoid_prefix
                }
                self.cbpsu_list.addItem(text, data)

            QMessageBox.information(self, "Loaded", "PSU list loaded.")
            if self.cbpsu_list.count() > 0:
                self.update_selected_psu_paths()

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # =========================================================
    # Project Abbreviation / Naming
    # =========================================================
    @staticmethod
    def normalize_project_abbreviation(value):
        value = str(value or "").strip().upper()
        value = re.sub(r"[^A-Z0-9]+", "_", value)
        value = re.sub(r"_+", "_", value).strip("_")
        return value

    def current_project_abbreviation(self, required=True):
        code = self.normalize_project_abbreviation(
            self.project_abbreviation.text()
        )

        if required and not code:
            raise ValueError(
                "Enter a Project Abbreviation first. "
                "Examples: LFS, GATS, HSDV."
            )

        return code

    def on_project_abbreviation_changed(self):
        raw = self.project_abbreviation.text()
        upper = raw.upper()

        if raw != upper:
            cursor = self.project_abbreviation.cursorPosition()
            self.project_abbreviation.blockSignals(True)
            self.project_abbreviation.setText(upper)
            self.project_abbreviation.setCursorPosition(
                min(cursor, len(upper))
            )
            self.project_abbreviation.blockSignals(False)

        if self.cbpsu_list.currentIndex() >= 0:
            self.update_selected_psu_paths()

    def build_project_key(self, year, round_number):
        """
        LFS is monthly:
            2026_JAN_LFS
            2026_FEB_LFS

        Other projects do not include LFS round/month:
            2026_GATS
            2026_HSDV
        """
        project_code = self.current_project_abbreviation()
        year = int(year)

        if project_code == "LFS":
            round_number = int(round_number)

            if not 1 <= round_number <= 12:
                raise ValueError(
                    f"Invalid LFS round/month: {round_number}. "
                    "Expected a value from 1 to 12."
                )

            month_abbr = self.month_name(round_number)[:3]
            return f"{year}_{month_abbr}_{project_code}"

        return f"{year}_{project_code}"

    def build_project_rollout_root(
        self,
        province,
        year,
        round_number,
    ):
        """
        LFS:
            C:/PSA-GIS/ROMBLON/GEOFASU/2026/JANUARY

        Other projects:
            C:/PSA-GIS/ROMBLON/GEOFASU/2026/GATS
            C:/PSA-GIS/ROMBLON/GEOFASU/2026/HSDV
        """
        project_code = self.current_project_abbreviation()
        province = str(
            province or "PROVINCE"
        ).strip().upper()
        year = int(year)

        root = os.path.join(
            "C:/PSA-GIS",
            province,
            "GEOFASU",
            str(year),
        )

        if project_code == "LFS":
            round_number = int(round_number)

            if not 1 <= round_number <= 12:
                raise ValueError(
                    f"Invalid LFS round/month: {round_number}."
                )

            return os.path.join(
                root,
                self.month_name(round_number),
            )

        return os.path.join(
            root,
            project_code,
        )

    # =========================================================
    # Selected PSU Path Builders
    # =========================================================
    def update_selected_psu_paths(self):
        idx = self.cbpsu_list.currentIndex()

        if idx < 0:
            self.output_path.clear()
            self.qfield_package_path.clear()
            return

        data = self.cbpsu_list.itemData(idx) or {}

        project_code = self.current_project_abbreviation(
            required=False
        ) or "PROJECT"

        year = int(data.get("Year", 0))
        rnd = int(data.get("Round", 0))
        psu_number = data.get("PSU_number")
        province = str(
            data.get("Prov_name", "PROVINCE")
        ).strip().upper()

        try:
            rollout_root = self.build_project_rollout_root(
                province=province,
                year=year,
                round_number=rnd,
            )
        except Exception:
            rollout_root = os.path.join(
                "C:/PSA-GIS",
                province,
                "GEOFASU",
                str(year),
                project_code,
            )

        self.output_path.setText(
            os.path.join(
                rollout_root,
                f"PSU_{psu_number}",
            )
        )

        self.qfield_package_path.setText(
            os.path.join(
                rollout_root,
                "QField Packages",
                f"PSU_{psu_number}",
            )
        )

        self.tblQFieldLayers.setRowCount(0)
        self._last_qfield_inspection = None
        self.pbPackageQField.setEnabled(False)

        try:
            project_key = self.build_project_key(
                year=year,
                round_number=rnd,
            )
        except Exception:
            project_key = project_code

        self.lblQFieldStatus.setText(
            f"{project_key} selected. Generate the PSU project, "
            "then inspect it for QField package readiness."
        )

    def update_default_output_path(self):
        """Backward-compatible alias."""
        self.update_selected_psu_paths()

    # =========================================================
    # QField Milestone 1 — Project Inspection
    # =========================================================
    def inspect_qfield_project(self):
        package_folder = self.qfield_package_path.text().strip()

        if self.cbpsu_list.currentIndex() < 0:
            QMessageBox.warning(
                self,
                "No PSU Selected",
                "Select a PSU before inspecting the current project."
            )
            return

        if not package_folder:
            QMessageBox.warning(
                self,
                "Missing Package Destination",
                "The QField package destination could not be determined."
            )
            return

        try:
            result = inspect_current_project(
                package_folder=package_folder,
            )

            self.tblQFieldLayers.setRowCount(
                len(result.layer_results)
            )

            error_items = []
            warning_items = []

            for row_index, item in enumerate(result.layer_results):
                values = [
                    item.layer_name,
                    item.layer_type,
                    item.provider,
                    item.proposed_action,
                    item.status,
                ]

                if item.status == "Error":
                    row_background = QColor("#fee2e2")
                    row_foreground = QColor("#991b1b")
                    error_items.append(
                        f"{item.layer_name}: {item.message}"
                    )

                elif item.status == "Warning":
                    row_background = QColor("#fef3c7")
                    row_foreground = QColor("#92400e")
                    warning_items.append(
                        f"{item.layer_name}: {item.message}"
                    )

                else:
                    row_background = QColor("#ecfdf5")
                    row_foreground = QColor("#166534")

                for column_index, value in enumerate(values):
                    table_item = QTableWidgetItem(str(value))
                    table_item.setToolTip(item.message)

                    table_item.setBackground(
                        QBrush(row_background)
                    )
                    table_item.setForeground(
                        QBrush(row_foreground)
                    )

                    self.tblQFieldLayers.setItem(
                        row_index,
                        column_index,
                        table_item
                    )

            self.tblQFieldLayers.resizeRowsToContents()

            self._last_qfield_inspection = result
            self.pbPackageQField.setEnabled(
                result.error_count == 0
            )

            if result.error_count > 0:
                self.lblQFieldStatus.setText(
                    "Project inspection failed. "
                    f"{result.error_count} blocking error(s) and "
                    f"{result.warning_count} warning(s) were found. "
                    "Resolve all red rows before packaging."
                )

            elif result.warning_count > 0:
                self.lblQFieldStatus.setText(
                    "Project inspection passed with warnings. "
                    f"{result.warning_count} warning(s) should be reviewed "
                    "before production packaging."
                )

            else:
                self.lblQFieldStatus.setText(
                    "Project inspection passed. "
                    f"{result.ready_count} item(s) are ready for "
                    "the proposed QField package actions."
                )

            message_parts = [
                "Inspection completed.",
                "",
                f"Ready: {result.ready_count}",
                f"Warnings: {result.warning_count}",
                f"Errors: {result.error_count}",
            ]

            if error_items:
                message_parts.extend([
                    "",
                    "BLOCKING ERRORS:",
                ])

                for index, message in enumerate(
                    error_items,
                    start=1,
                ):
                    message_parts.append(
                        f"{index}. {message}"
                    )

            if warning_items:
                message_parts.extend([
                    "",
                    "WARNINGS:",
                ])

                for index, message in enumerate(
                    warning_items,
                    start=1,
                ):
                    message_parts.append(
                        f"{index}. {message}"
                    )

            if result.error_count > 0:
                message_parts.extend([
                    "",
                    "Resolve the blocking errors shown above before "
                    "QField packaging is enabled."
                ])

                QMessageBox.warning(
                    self,
                    "QField Project Inspection",
                    "\n".join(message_parts)
                )

            else:
                message_parts.extend([
                    "",
                    "No blocking errors were found. Package for QField is now enabled."
                ])

                QMessageBox.information(
                    self,
                    "QField Project Inspection",
                    "\n".join(message_parts)
                )

        except Exception as e:
            QMessageBox.critical(
                self,
                "QField Inspection Error",
                str(e)
            )


    # =========================================================
    # QField Milestone 2 — Package Current Project
    # =========================================================
    def package_qfield_project(self):
        package_folder = self.qfield_package_path.text().strip()

        if self.cbpsu_list.currentIndex() < 0:
            QMessageBox.warning(
                self,
                "No PSU Selected",
                "Select a PSU before creating a QField package."
            )
            return

        current_project_file = str(
            QgsProject.instance().fileName() or ""
        ).strip()

        if (
            not current_project_file
            or not os.path.isfile(current_project_file)
        ):
            QMessageBox.warning(
                self,
                "Project Not Saved",
                "Generate or save the current PSU project before packaging."
            )
            return

        if not package_folder:
            QMessageBox.warning(
                self,
                "Missing Package Destination",
                "Select a PSU first so the QField package destination "
                "can be determined."
            )
            return

        # Always run the latest inspection before packaging.
        try:
            inspection = inspect_current_project(
                package_folder=package_folder,
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "QField Inspection Error",
                str(e)
            )
            return

        if (
            self.chkQFieldValidate.isChecked()
            and inspection.error_count > 0
        ):
            self._last_qfield_inspection = inspection
            self.pbPackageQField.setEnabled(False)

            QMessageBox.warning(
                self,
                "QField Packaging Blocked",
                "The current project has blocking QField package errors.\n\n"
                "Run Inspect Current Project and resolve all red rows first."
            )
            return

        package_path = Path(package_folder)

        if (
            package_path.exists()
            and any(package_path.iterdir())
        ):
            answer = QMessageBox.question(
                self,
                "Replace Existing QField Package",
                "The QField package folder already contains files:\n\n"
                f"{package_folder}\n\n"
                "Delete the existing package contents and create a new "
                "package for this PSU?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )

            if answer != QMessageBox.Yes:
                return

            try:
                shutil.rmtree(package_path)
            except OSError as exc:
                QMessageBox.critical(
                    self,
                    "Cannot Replace QField Package",
                    "The existing package folder could not be removed.\n\n"
                    f"{package_folder}\n\n"
                    f"Error:\n{exc}"
                )
                return

        package_path.mkdir(
            parents=True,
            exist_ok=True,
        )

        QApplication.setOverrideCursor(
            Qt.WaitCursor
        )

        progress = QProgressDialog(
            "Creating QField package...",
            "",
            0,
            0,
            self,
        )
        progress.setWindowTitle(
            "GEOFASU — QField Package"
        )
        progress.setWindowModality(
            Qt.WindowModal
        )
        progress.setCancelButton(None)
        progress.setMinimumDuration(0)
        progress.show()

        try:
            canvas_extent = None

            if (
                self.cmbQFieldExtent.currentText()
                == "Current map canvas"
            ):
                from qgis.utils import iface

                canvas_extent = (
                    iface.mapCanvas().extent()
                )

            result = package_current_project(
                package_folder=package_folder,
                extent_mode=(
                    self.cmbQFieldExtent
                    .currentText()
                ),
                canvas_extent=canvas_extent,
                copy_styles_resources=(
                    self.chkQFieldStyles
                    .isChecked()
                ),
                preserve_snapping=(
                    self.chkQFieldSnapping
                    .isChecked()
                ),
                relative_paths=(
                    self.chkQFieldRelativePaths
                    .isChecked()
                ),
            )

            self.lblQFieldStatus.setText(
                "QField package created successfully. "
                f"{len(result.packaged_layers)} layer(s) packaged."
            )

            QMessageBox.information(
                self,
                "QField Package Complete",
                "The QField cable package was created successfully.\n\n"
                f"Package folder:\n{result.package_folder}\n\n"
                f"QField project:\n{result.packaged_project_file}\n\n"
                f"Offline data:\n{result.offline_database}\n\n"
                f"Manifest:\n{result.manifest_file}\n\n"
                "Copy the complete PSU package folder to the QField device."
            )

            # The packager reloads the original desktop project.
            # Re-inspect it so the UI reflects the restored project.
            self.inspect_qfield_project()

        except Exception as e:
            QMessageBox.critical(
                self,
                "QField Packaging Error",
                str(e)
            )

        finally:
            progress.close()
            QApplication.restoreOverrideCursor()

    # =========================================================
    # Remove Temporary Processing Layers
    # =========================================================
    @staticmethod
    def remove_temporary_processing_layers():
        """
        Remove temporary memory layers created by processing algorithms.

        GEOFASU's final project should contain only persistent/intentional
        layers such as:
            - the generated LFS GeoPackage layer
            - BARANGAY BOUNDARY
            - BASEMAP

        Processing may temporarily register generic memory layers named
        "output". These must not be saved into the PSU project or considered
        during QField packaging.
        """
        project = QgsProject.instance()

        layer_ids_to_remove = []

        for layer_id, layer in project.mapLayers().items():
            try:
                provider = str(layer.providerType() or "").casefold()
            except Exception:
                provider = ""

            try:
                layer_name = str(layer.name() or "").strip().casefold()
            except Exception:
                layer_name = ""

            # Only remove generic temporary processing outputs.
            # Do not remove intentionally named memory layers.
            if provider == "memory" and layer_name == "output":
                layer_ids_to_remove.append(layer_id)

        if layer_ids_to_remove:
            project.removeMapLayers(layer_ids_to_remove)

        return len(layer_ids_to_remove)

    # =========================================================
    # Remove Temporary Processing Layers
    # =========================================================
    @staticmethod
    def remove_temporary_processing_layers():
        """
        Remove generic memory layers named "output" created by processing.
        """
        project = QgsProject.instance()
        layer_ids_to_remove = []

        for layer_id, layer in project.mapLayers().items():
            try:
                provider = str(
                    layer.providerType() or ""
                ).casefold()
            except Exception:
                provider = ""

            try:
                layer_name = str(
                    layer.name() or ""
                ).strip().casefold()
            except Exception:
                layer_name = ""

            if (
                provider == "memory"
                and layer_name == "output"
            ):
                layer_ids_to_remove.append(
                    layer_id
                )

        if layer_ids_to_remove:
            project.removeMapLayers(
                layer_ids_to_remove
            )

        return len(layer_ids_to_remove)

    # =========================================================
    # Generate Geometry
    # =========================================================
    def generate_geometry(self):
        try:
            project_code = self.current_project_abbreviation()
        except ValueError as exc:
            QMessageBox.warning(
                self,
                "Project Abbreviation Required",
                str(exc)
            )
            return

        path = self.ssu_list_path.text()
        if not path:
            QMessageBox.warning(self, "Missing File", "Select Excel file first.")
            return

        output_folder = self.output_path.text()
        if not output_folder:
            QMessageBox.warning(self, "Missing Output Folder", "Select output folder first.")
            return
        os.makedirs(output_folder, exist_ok=True)

        QApplication.setOverrideCursor(Qt.WaitCursor)
        progress = QProgressDialog("Processing...", "Cancel", 0, 0, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        try:
            feedback = QgsProcessingFeedback()
            num_pat = re.compile(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?')
            main_layer = QgsVectorLayer("Point?crs=EPSG:4326", "SSU + Replacement (temp)", "memory")
            main_dp = main_layer.dataProvider()

            primary_sheets = ['Sample SSU', 'Selected Samples', 'AONCR-SAMPLE HH']
            replacement_sheet = 'Replacement SSU'
            layers = []
            wkt_fields = {}

            for name in primary_sheets + [replacement_sheet]:
                uri = f'{path}|layername={name}'
                lyr = QgsVectorLayer(uri, name.lower().replace(' ', '_'), 'ogr')
                if not lyr.isValid():
                    continue
                wkt_field = next((f.name() for f in lyr.fields() if 'wkt' in f.name().lower()), None)
                if not wkt_field:
                    continue
                layers.append(lyr)
                wkt_fields[lyr] = wkt_field

            if not layers:
                QMessageBox.warning(self, "No data", "No valid sheets found.")
                return

            main_fields = layers[0].fields()
            main_dp.addAttributes(main_fields)
            main_layer.updateFields()

            for lyr in layers:
                for feat in lyr.getFeatures():
                    wkt = feat[wkt_fields[lyr]]
                    geom = QgsGeometry.fromWkt(wkt)
                    if geom.isNull():
                        nums = num_pat.findall(str(wkt))
                        if len(nums) >= 2:
                            geom = QgsGeometry.fromPointXY(QgsPointXY(float(nums[0]), float(nums[1])))
                    if geom.isNull():
                        continue
                    nf = QgsFeature(main_fields)
                    nf.setGeometry(geom)
                    nf.setAttributes(feat.attributes())
                    main_dp.addFeature(nf)

            main_layer.updateExtents()

            idx = self.cbpsu_list.currentIndex()
            if idx < 0:
                QMessageBox.warning(self, "No PSU selected", "Select a PSU.")
                return

            psu_data = self.cbpsu_list.itemData(idx)
            psu_number = psu_data["PSU_number"]
            geoid_prefix = psu_data.get("Geoid_prefix")

            filtered_layer = extract_by_psu(main_layer, psu_number, feedback)
            refactored_layer = refactor_psu_layer(filtered_layer, context=None, feedback=feedback)

            year = int(psu_data["Year"])
            rnd = int(psu_data["Round"])
            rep = psu_data.get("Replicate_Number")

            try:
                rep = int(rep)
            except Exception:
                rep = str(rep)

            prov_name = str(
                psu_data["Prov_name"]
            ).strip().upper()

            project_key = self.build_project_key(
                year=year,
                round_number=rnd,
            )

            base_name = (
                f"{project_key}_{prov_name}"
                f"_SELECTED_SSU_R{rep}_PSU_{psu_number}"
            )

            # --- Save project GeoPackage ---
            filename = f"{base_name}.gpkg"
            output_file = os.path.join(output_folder, filename)

            QgsVectorFileWriter.writeAsVectorFormat(
                refactored_layer, output_file, "UTF-8", QgsCoordinateReferenceSystem("EPSG:4326"), "GPKG"
            )

            # --- Load layers to keep references ---
            self.lfs_layer = QgsVectorLayer(output_file, filename, "ogr")
            self.lfs_layer.setCrs(QgsCoordinateReferenceSystem("EPSG:4326"))
            self.lfs_layer.setName(base_name)

            # Snapping config
            snap_cfg = QgsProject.instance().snappingConfig()
            snap_cfg.setEnabled(True)
            snap_cfg.setType(QgsSnappingConfig.Vertex)
            snap_cfg.setMode(QgsSnappingConfig.AllLayers)
            snap_cfg.setTolerance(12)
            snap_cfg.setUnits(QgsTolerance.Pixels)
            snap_cfg.setIntersectionSnapping(False)
            QgsProject.instance().setSnappingConfig(snap_cfg)

            # =================================================
            # Load Barangay layer
            # =================================================
            bgy_layer = load_barangay_layer_smart(self.lfs_layer, geoid_prefix=geoid_prefix, output_folder=output_folder)
            clipped_raster = None
            if bgy_layer:
                self.bgy_layer = bgy_layer
                self.bgy_layer.setName("BARANGAY BOUNDARY")
                self.bgy_layer.setCrs(QgsCoordinateReferenceSystem("EPSG:4326"))
                try:
                    clipped_raster = clip_raster_by_bgy_memory(bgy_layer, geoid_prefix, output_folder)
                    self.clipped_raster = clipped_raster
                except Exception as e:
                    QMessageBox.warning(self, "Raster Clip", f"Could not clip raster:\n{str(e)}")

            # =================================================
            # Apply Styles
            # =================================================
            try:
                if self.lfs_layer.isValid():
                    load_style(self.lfs_layer, "samples_rep_layer.qml")
                if self.bgy_layer and self.bgy_layer.isValid():
                    load_style(self.bgy_layer, "bgy_boundary.qml")
            except Exception as e:
                QMessageBox.warning(self, "Style Warning", f"Could not apply style:\n{str(e)}")

            # =================================================
            # Layer Groups
            # =================================================
            root = QgsProject.instance().layerTreeRoot()
            project_group_name = f"{project_code} Layers"
            lfs_group = (
                root.findGroup(project_group_name)
                or root.addGroup(project_group_name)
            )
            base_group = root.findGroup("Base Layer") or root.addGroup("Base Layer")
            basemap_group = root.findGroup("Basemap") or root.addGroup("Basemap")

            if self.lfs_layer.isValid():
                QgsProject.instance().addMapLayer(self.lfs_layer, False)
                node = lfs_group.addLayer(self.lfs_layer)
                node.setExpanded(False)

            if self.bgy_layer and self.bgy_layer.isValid():
                QgsProject.instance().addMapLayer(self.bgy_layer, False)
                node = base_group.addLayer(self.bgy_layer)
                node.setExpanded(False)

            if self.clipped_raster and self.clipped_raster.isValid():
                QgsProject.instance().addMapLayer(self.clipped_raster, False)
                node = basemap_group.addLayer(self.clipped_raster)
                node.setItemVisibilityChecked(False)
                node.setExpanded(False)

            # --- Zoom to outputs ---
            from qgis.utils import iface
            combined_extent = self.lfs_layer.extent()
            if self.bgy_layer:
                combined_extent.combineExtentWith(self.bgy_layer.extent())
            if self.clipped_raster:
                combined_extent.combineExtentWith(self.clipped_raster.extent())
            iface.mapCanvas().setExtent(combined_extent)
            iface.mapCanvas().refresh()

            # -------------------------------------------------
            # Remove temporary processing memory layers
            # -------------------------------------------------
            removed_temp_layers = self.remove_temporary_processing_layers()

            # -------------------------------------------------
            # Remove generic temporary processing layers
            # -------------------------------------------------
            removed_temp_layers = self.remove_temporary_processing_layers()

            # --- Save QGIS Project ---
            project_filename = f"{base_name}.qgs"
            self.generated_project_path = os.path.join(output_folder, project_filename)
            QgsProject.instance().setCrs(QgsCoordinateReferenceSystem("EPSG:4326"))
            QgsProject.instance().write(self.generated_project_path)

            self.update_selected_psu_paths()

            # Defensive cleanup in case a processing provider registered
            # another generic memory "output" layer after project writing.
            self.remove_temporary_processing_layers()

            self.inspect_qfield_project()

            QMessageBox.information(
                self,
                "Done",
                f"{project_key} PSU geometry and QGIS project were generated successfully.\n\n"
                f"GeoPackage:\n{output_file}\n\n"
                f"QGIS project:\n{self.generated_project_path}\n\n"
                "QField package readiness was also inspected."
                + (
                    f"\n\nTemporary processing layers removed: "
                    f"{removed_temp_layers}"
                    if removed_temp_layers
                    else ""
                )
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
        finally:
            progress.close()
            QApplication.restoreOverrideCursor()